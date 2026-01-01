"""
Main processor that orchestrates the complete receipt processing pipeline.
"""
import time
from pathlib import Path
from typing import List, Optional, Union
import pandas as pd

from core.preprocessing import ImagePreprocessor
from core.ocr_engine import OCREngine
from core.extractors import FieldExtractor
from core.hybrid_extractor import HybridFieldExtractor
from core.validators import DataValidator
from core.models import (
    ReceiptResult,
    BatchResult,
    ReceiptType
)
from config.settings import OUTPUT_DIR, EXPORT_SETTINGS


class ReceiptProcessor:
    """
    Main processor for receipt OCR and data extraction.
    Orchestrates preprocessing, OCR, extraction, and validation.
    """

    def __init__(
        self,
        ocr_engine: str = "auto",
        enable_preprocessing: bool = True,
        enable_validation: bool = True,
        use_hybrid: bool = True
    ):
        """
        Initialize the receipt processor.

        Args:
            ocr_engine: OCR engine to use ('auto', 'easyocr', or 'tesseract')
            enable_preprocessing: Whether to preprocess images
            enable_validation: Whether to validate extracted data
            use_hybrid: Whether to use hybrid extraction (recommended)
        """
        print("Initializing Receipt Processor...")

        self.enable_preprocessing = enable_preprocessing
        self.enable_validation = enable_validation
        self.use_hybrid = use_hybrid

        # Initialize components
        self.preprocessor = ImagePreprocessor()
        print("✓ Preprocessor initialized")

        self.ocr_engine = OCREngine(engine=ocr_engine)
        print("✓ OCR Engine initialized")

        if use_hybrid:
            self.extractor = HybridFieldExtractor(self.ocr_engine)
            print("✓ Hybrid Extractor initialized (Smart Slicing + Box Detection)")
        else:
            self.extractor = FieldExtractor(self.ocr_engine)
            print("✓ Field Extractor initialized")

        self.validator = DataValidator()
        print("✓ Validator initialized")

        print("Receipt Processor ready!\n")

    def process_receipt(
        self,
        image_path: Union[str, Path],
        save_debug: bool = False
    ) -> ReceiptResult:
        """
        Process a single receipt image.

        Args:
            image_path: Path to the receipt image
            save_debug: Whether to save debug images

        Returns:
            ReceiptResult object
        """
        start_time = time.time()
        image_path = Path(image_path)

        print(f"Processing: {image_path.name}")

        try:
            # Step 1: Preprocess image and detect type
            if self.enable_preprocessing:
                original, processed, receipt_type = self.preprocessor.preprocess_for_ocr(
                    image_path,
                    enhance=True,
                    deskew=False,  # Mobile screenshots don't need deskewing
                    aggressive=False  # Use gentle enhancement
                )
                print(f"  ✓ Preprocessed (Type: {receipt_type.value})")
            else:
                original = self.preprocessor.load_image(image_path)
                processed = original
                receipt_type = self.preprocessor.detect_receipt_type(original)
                print(f"  ✓ Loaded (Type: {receipt_type.value})")

            # Save debug image if requested
            if save_debug:
                debug_path = OUTPUT_DIR / f"debug_{image_path.stem}_processed.png"
                self.preprocessor.save_debug_image(processed, debug_path)
                print(f"  ✓ Debug image saved: {debug_path.name}")

            # Step 2: Extract fields
            if self.use_hybrid:
                extracted_data = self.extractor.extract_fields(
                    processed,
                    receipt_type.value,
                    debug=save_debug
                )
            else:
                extracted_data = self.extractor.extract_fields(processed, receipt_type)
            print(f"  ✓ Fields extracted")

            # Step 3: Validate data
            if self.enable_validation:
                result = self.validator.validate_receipt(
                    extracted_data,
                    receipt_type,
                    image_path.name
                )
                print(f"  ✓ Validated (Confidence: {result.overall_confidence:.2%}, "
                      f"Flagged: {result.flagged})")
            else:
                # Create result without validation
                result = ReceiptResult(
                    filename=image_path.name,
                    receipt_type=receipt_type,
                    data=extracted_data,
                    overall_confidence=0.0,
                    flagged=False
                )

            # Add processing time
            processing_time = time.time() - start_time
            result.processing_time = processing_time

            print(f"  ✓ Completed in {processing_time:.2f}s\n")

            return result

        except Exception as e:
            print(f"  ✗ Error: {str(e)}\n")
            # Return a failed result
            from core.models import TransactionData, ValidationIssue

            return ReceiptResult(
                filename=image_path.name,
                receipt_type=ReceiptType.UNKNOWN,
                data=TransactionData(),
                overall_confidence=0.0,
                flagged=True,
                validation_issues=[ValidationIssue(
                    field="processing",
                    issue_type="error",
                    message=f"Processing failed: {str(e)}",
                    severity="error"
                )],
                processing_time=time.time() - start_time
            )

    def process_batch(
        self,
        image_paths: List[Union[str, Path]],
        save_debug: bool = False
    ) -> BatchResult:
        """
        Process multiple receipt images.

        Args:
            image_paths: List of paths to receipt images
            save_debug: Whether to save debug images

        Returns:
            BatchResult object
        """
        start_time = time.time()

        print(f"Processing batch of {len(image_paths)} receipts...\n")
        print("=" * 60)

        results = []
        for image_path in image_paths:
            result = self.process_receipt(image_path, save_debug=save_debug)
            results.append(result)

        # Calculate statistics
        successful = len([r for r in results if not r.flagged and r.receipt_type != ReceiptType.UNKNOWN])
        flagged = len([r for r in results if r.flagged])
        failed = len([r for r in results if r.receipt_type == ReceiptType.UNKNOWN])

        total_time = time.time() - start_time

        print("=" * 60)
        print(f"\nBatch processing complete!")
        print(f"  Total: {len(results)}")
        print(f"  Successful: {successful}")
        print(f"  Flagged for review: {flagged}")
        print(f"  Failed: {failed}")
        print(f"  Total time: {total_time:.2f}s")
        print(f"  Average time per receipt: {total_time/len(results):.2f}s\n")

        return BatchResult(
            total_processed=len(results),
            successful=successful,
            flagged=flagged,
            failed=failed,
            results=results,
            processing_time=total_time
        )

    def process_directory(
        self,
        directory_path: Union[str, Path],
        pattern: str = "*.*",
        save_debug: bool = False
    ) -> BatchResult:
        """
        Process all receipt images in a directory.

        Args:
            directory_path: Path to directory containing receipts
            pattern: File pattern to match (default: *.* for all files)
            save_debug: Whether to save debug images

        Returns:
            BatchResult object
        """
        directory = Path(directory_path)

        if not directory.exists():
            raise ValueError(f"Directory does not exist: {directory}")

        # Find all matching images (filter by extension)
        valid_extensions = {'.png', '.jpg', '.jpeg', '.PNG', '.JPG', '.JPEG'}
        all_files = sorted(directory.glob(pattern))
        image_files = [f for f in all_files if f.suffix in valid_extensions]

        if not image_files:
            print(f"No images found matching pattern '{pattern}' in {directory}")
            return BatchResult(
                total_processed=0,
                successful=0,
                flagged=0,
                failed=0,
                results=[],
                processing_time=0.0
            )

        print(f"Found {len(image_files)} images in {directory}\n")

        return self.process_batch(image_files, save_debug=save_debug)

    def export_to_excel(
        self,
        batch_result: BatchResult,
        output_path: Optional[Union[str, Path]] = None,
        include_flagged: bool = True
    ) -> Path:
        """
        Export batch results to Excel file.

        Args:
            batch_result: BatchResult from processing
            output_path: Path for output file (default: auto-generated)
            include_flagged: Whether to include flagged receipts

        Returns:
            Path to the created Excel file
        """
        if output_path is None:
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            output_path = OUTPUT_DIR / f"receipts_{timestamp}.xlsx"
        else:
            output_path = Path(output_path)

        # Convert results to dictionaries
        results_to_export = batch_result.results
        if not include_flagged:
            results_to_export = [r for r in results_to_export if not r.flagged]

        data_dicts = [result.to_dict() for result in results_to_export]

        # Create DataFrame
        df = pd.DataFrame(data_dicts)

        # Reorder columns according to settings
        column_order = EXPORT_SETTINGS['excel_columns']
        df = df[[col for col in column_order if col in df.columns]]

        # Export to Excel with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Receipts', index=False)

            # Get worksheet for formatting
            worksheet = writer.sheets['Receipts']

            # Auto-adjust column widths
            for idx, col in enumerate(df.columns, 1):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(col)
                ) + 2
                worksheet.column_dimensions[chr(64 + idx)].width = min(max_length, 50)

            # Highlight flagged rows in yellow
            from openpyxl.styles import PatternFill
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            for row_idx, row in enumerate(df.itertuples(), start=2):
                if row.flagged:
                    for col_idx in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = yellow_fill

        print(f"✓ Exported {len(data_dicts)} receipts to: {output_path}")

        return output_path

    def export_flagged_report(
        self,
        batch_result: BatchResult,
        output_path: Optional[Union[str, Path]] = None
    ) -> Optional[Path]:
        """
        Export a detailed report of flagged receipts.

        Args:
            batch_result: BatchResult from processing
            output_path: Path for output file (default: auto-generated)

        Returns:
            Path to the created file, or None if no flagged receipts
        """
        flagged = batch_result.get_flagged_receipts()

        if not flagged:
            print("No flagged receipts to report")
            return None

        if output_path is None:
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            output_path = OUTPUT_DIR / f"flagged_receipts_{timestamp}.txt"
        else:
            output_path = Path(output_path)

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("FLAGGED RECEIPTS REPORT\n")
            f.write("=" * 60 + "\n\n")
            f.write(f"Total flagged: {len(flagged)}\n\n")

            for idx, result in enumerate(flagged, 1):
                f.write(f"{idx}. {result.filename}\n")
                f.write(f"   Receipt Type: {result.receipt_type.value}\n")
                f.write(f"   Confidence: {result.overall_confidence:.2%}\n")
                f.write(f"   Issues:\n")

                for issue in result.validation_issues:
                    f.write(f"     - [{issue.severity.upper()}] {issue.field}: {issue.message}\n")

                f.write(f"   Extracted Data:\n")
                for field_name in ['transaction_id', 'datetime', 'from_account',
                                   'to_account', 'receiver_name', 'amount']:
                    value = result.data.get_field_value(field_name)
                    conf = result.data.get_field_confidence(field_name)
                    if value:
                        f.write(f"     {field_name}: {value} (confidence: {conf:.2%})\n")

                f.write("\n" + "-" * 60 + "\n\n")

        print(f"✓ Flagged receipts report saved to: {output_path}")

        return output_path


# Convenience function
def create_processor(ocr_engine: str = "auto") -> ReceiptProcessor:
    """
    Factory function to create a receipt processor.

    Args:
        ocr_engine: OCR engine to use ('auto', 'easyocr', or 'tesseract')

    Returns:
        ReceiptProcessor instance
    """
    return ReceiptProcessor(ocr_engine=ocr_engine)